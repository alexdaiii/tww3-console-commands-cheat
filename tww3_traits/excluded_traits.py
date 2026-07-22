#!/usr/bin/env python3
"""
excluded_traits.py - The inverse of curate_traits.py.

Walks EVERY character trait and reports the ones that did NOT make it into the
power set, with the exact reason each was rejected. Reuses curate_traits'
filter helpers so the two scripts can never drift apart.

Reason is the FIRST rule (in curate_traits' own order) that dropped the trait:
  hidden -> internal (dummy) -> DLC-locked -> no effects (flavour) ->
  self-debuff -> corruption -> bundled downside -> off-target.

Output: excluded_traits.xlsx  (filter the "Reason" / "Game" columns in Excel).
"""
from __future__ import annotations
import os
import join_traits as J
import curate_traits as C

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(HERE, "excluded_traits.xlsx")


def classify(rep_eff):
    """Return (reason, bad_effect) for the effect-level filters, or (None, '')
    if the trait passes them all (i.e. it is NOT rejected on effects)."""
    for rendered, scope, val in rep_eff:
        low = rendered.lower()
        sc = (scope or "").lower()
        if ("enemy" not in low and "enemy" not in sc and C._neg_value(val)
                and any(s in low for s in C.HARM_STATS)):
            return ("self-debuff (negative combat stat)", rendered)
        try:
            if "cooldown" in low and float(val) > 0:
                return ("self-debuff (increased cooldown)", rendered)
        except (TypeError, ValueError):
            pass
    for rendered, scope, val in rep_eff:
        if C.CORRUPTION_KW in rendered.lower():
            return ("corruption effect", rendered)
    for rendered, scope, val in rep_eff:
        if C.is_unwanted(rendered, scope):
            return ("bundled downside (plague / -public order / -diplo / WoM drain)",
                    rendered)
    # bucket every effect; if none land in a target area it's off-target
    buckets = {b: [] for b in ("econ", "survive", "army", "attack", "diplo", "other")}
    for rendered, scope, _ in rep_eff:
        buckets[C.bucket(rendered, scope)].append(rendered)
    if not (buckets["econ"] or buckets["survive"] or buckets["army"]
            or buckets["attack"] or buckets["diplo"]):
        return ("off-target (only flavour/'other' effects)", "")
    return (None, "")


def main():
    tl_loc = J.read_loc("trait_levels_loc")
    eff_loc = J.read_loc("effects_loc")

    def name_of(k):
        return J.clean_text(tl_loc.get(f"character_trait_levels_onscreen_name_{k}", ""))

    raw_by_level: dict[str, list[tuple[str, str, str]]] = {}
    for r in J.read_tsv("trait_level_effects"):
        lk = r.get("trait_level", "")
        if not lk:
            continue
        rendered = J.render_effect(r.get("effect", ""), r.get("value", ""), eff_loc)
        raw_by_level.setdefault(lk, []).append(
            (rendered, r.get("effect_scope", ""), r.get("value", "")))

    levels_by_trait: dict[str, list[dict]] = {}
    for r in J.read_tsv("character_trait_levels"):
        base = r.get("trait", "")
        lk = r.get("key", "")
        try:
            lvl = int(r.get("level", "0") or 0)
        except ValueError:
            lvl = 0
        levels_by_trait.setdefault(base, []).append(
            {"lk": lk, "lvl": lvl, "eff": raw_by_level.get(lk, [])})

    rejects = []
    for r in J.read_tsv("character_traits"):
        key = r.get("key", "")
        parts = key.split("_")
        dlc_token = parts[1] if len(parts) > 1 else ""
        lvls = levels_by_trait.get(key, [])
        rep = max(lvls, key=lambda l: (l["lvl"], bool(l["eff"]), l["lk"] == key),
                  default=None)
        rep_eff = rep["eff"] if rep else []
        # effects text for the report (best available: rep, else any level)
        show_eff = rep_eff
        if not show_eff:
            for l in lvls:
                if l["eff"]:
                    show_eff = l["eff"]
                    break

        # --- reasons, in curate_traits' own evaluation order ---
        if (r.get("hidden", "") or "").lower() == "true":
            reason, bad = "hidden trait", ""
        elif "dummy" in key:
            reason, bad = "internal placeholder (dummy)", ""
        elif dlc_token != "main":
            reason, bad = "DLC-locked (not base game)", ""
        elif not rep or not rep_eff:
            reason, bad = "no effects (flavour / personality trait)", ""
        elif key in C.FORCE_INCLUDE:
            continue   # explicitly promoted into the power set
        else:
            reason, bad = classify(rep_eff)
            if reason is None:
                continue   # this trait DID make the power set - skip

        rejects.append({
            "id": key,
            "name": name_of(rep["lk"] if rep else key) or name_of(key),
            "game": J.game_and_dlc(key)[0],
            "reason": reason,
            "trigger": bad,
            "effects": " | ".join(e[0] for e in show_eff),
        })

    rejects.sort(key=lambda x: (x["reason"], x["id"]))
    write_xlsx(rejects)
    print(f"Wrote {OUT_XLSX}  ({len(rejects)} excluded traits)")
    from collections import Counter
    for reason, n in Counter(x["reason"] for x in rejects).most_common():
        print(f"  {n:4d}  {reason}")


HEADER_FILL = PatternFill("solid", fgColor="7F1D1D")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_xlsx(rejects):
    wb = Workbook()
    ws = wb.active
    ws.title = "excluded_traits"
    columns = [
        ("reason", "Reason excluded"),
        ("id", "Trait ID"),
        ("name", "Name"),
        ("game", "Game"),
        ("trigger", "Triggering effect"),
        ("effects", "All effects"),
    ]
    ws.append([c[1] for c in columns])
    for i in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for x in rejects:
        ws.append([x[c[0]] for c in columns])

    widths = {"reason": 44, "id": 46, "name": 26, "game": 12,
              "trigger": 46, "effects": 80}
    for i, (k, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(k, 18)
    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, ws.max_row + 1):
        for k in ("name", "trigger", "effects"):
            i = [j for j, (kk, _) in enumerate(columns, start=1) if kk == k][0]
            ws.cell(row=r, column=i).alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    ws2 = wb.create_sheet("README")
    for row in [
        ["Excluded traits", ""],
        ["", ""],
        ["What this is", "Every character trait that did NOT make it into power_traits.xlsx, "
         "with the reason. Inverse of curate_traits.py; shares its filter logic."],
        ["Reason", "The first rule that dropped the trait. Filter this column to review a "
         "single category (e.g. 'off-target' = traits with only flavour effects you could "
         "still choose to add back)."],
        ["hidden / internal", "Engine-hidden or dummy placeholder traits - not grantable in a useful way."],
        ["DLC-locked", "Trait id is not wh_main / wh2_main / wh3_main. Works only with that DLC's data."],
        ["no effects", "Pure personality/flavour - no trait_level_effects rows at any level."],
        ["self-debuff", "Carries a negative combat/leadership stat aimed at your own side."],
        ["corruption", "Touches corruption (kept out by user rule; good/bad depends on faction)."],
        ["bundled downside", "Rides along a penalty: plague, -public order, -diplomacy, or a WoM reserve drain."],
        ["off-target", "Only 'other' effects - nothing in Survivability/Attack/Army/Economy/Diplomacy."],
        ["Source", "Local clone of https://github.com/Shazbot/WH3-Dump."],
    ]:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 110
    ws2["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws2.max_row + 1):
        ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
