#!/usr/bin/env python3
"""
list_effects.py  -  Dump the master list of EVERY effect key to all_effects.xlsx.

The effects table (db/effects_tables/data__.tsv) is the game's global registry of
every mechanical effect that traits, items (ancillaries), skills, technologies,
buildings and effect_bundles can reference. This is the vocabulary you draw from
when building any apply-Lua: an effect_bundle, trait, or item ultimately grants a
set of these keys with values.

One row per effect (~15k). Columns:
  effect                 the effect key (what junction tables point at)
  readable               loc template with the value placeholder shown as N / +N
  category               campaign | battle | ... (effects_tables.category)
  positive_is_good       is a positive value beneficial? (effects_tables)
  priority               sort/display priority
  has_loc                whether a human-readable description exists
  icon                   ui icon path

Run:  agent-venv/bin/python list_effects.py            -> writes ./all_effects.xlsx
      agent-venv/bin/python list_effects.py /path/to/WH3-Dump
"""
from __future__ import annotations
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import join_traits as J  # reuse read_tsv / read_loc / clean_text

# Register the extra tables on J's file map so J.read_tsv (which skips the
# CA #comment line) can read them too.
J.FILES["effects"] = "db/effects_tables/data__.tsv"
J.FILES["unit_abilities"] = "db/unit_abilities_tables/data__.tsv"
J.FILES["unit_abilities_loc"] = "text/db/unit_abilities__.loc.tsv"
# Direct effect -> single unit_ability bridge (cooldown mods, ability enablers).
J.FILES["eff_ability_jct"] = "db/effect_bonus_value_unit_ability_junctions_tables/data__.tsv"
# Mechanical detail for abilities: hand-authored UI bullets + phase numbers.
J.FILES["ui_effect_jct"] = "db/unit_abilities_to_additional_ui_effects_juncs_tables/data__.tsv"
J.FILES["ui_effect_loc"] = "text/db/unit_abilities_additional_ui_effects__.loc.tsv"
J.FILES["ab_phase_jct"] = "db/special_ability_to_special_ability_phase_junctions_tables/data__.tsv"
J.FILES["phases"] = "db/special_ability_phases_tables/data__.tsv"
J.FILES["phase_stats"] = "db/special_ability_phase_stat_effects_tables/data__.tsv"
J.FILES["phase_attrs"] = "db/special_ability_phase_attribute_effects_tables/data__.tsv"
# Cost / cooldown / uses for each ability (mana_cost = Winds of Magic).
J.FILES["unit_special_abilities"] = "db/unit_special_abilities_tables/data__.tsv"
OUT = os.path.join(J.HERE, "all_effects.xlsx")


def _num(v: str) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _clean_stat(stat: str) -> str:
    return stat.replace("stat_", "").replace("_", " ").strip()


def load_ability_cost() -> dict[str, dict]:
    """ability_key -> {wom, recharge, uses, passive} from unit_special_abilities."""
    out: dict[str, dict] = {}
    for r in J.read_tsv("unit_special_abilities"):
        k = r.get("key", "")
        if not k:
            continue
        out[k] = {
            "wom": r.get("mana_cost", ""),        # Winds of Magic cost (0 = free)
            "recharge": r.get("recharge_time", ""),
            "uses": r.get("num_uses", ""),        # -1 = unlimited
            "passive": r.get("passive", ""),
        }
    return out


def _fmt_wom(c: dict) -> str:
    if c.get("passive") == "true":
        return "free (passive)"
    return f"{_num(c['wom']):g} WoM" if _num(c.get("wom")) > 0 else "free"


def _fmt_cd(c: dict) -> str:
    f = _num(c.get("recharge"))
    return f"{f:g}s" if f > 0 else ""


def _fmt_uses(c: dict) -> str:
    try:
        n = int(float(c.get("uses") or 0))
    except ValueError:
        return ""
    return "unlimited" if n < 0 else str(n)


def load_ability_mechanics() -> dict[str, str]:
    """ability_key -> a compact 'what it actually does' string.

    Combines two in-game sources:
      * hand-authored UI bullets (unit_abilities_additional_ui_effects) — e.g.
        'Summons a unit of Ushabti', 'Effect increases with friends in range'.
      * special_ability_phases numbers — duration, damage, heal, barrier,
        resurrect, stat buffs and granted attributes, tagged by target scope.
    """
    # 1) UI bullets: ability -> [readable bullet, ...]
    ui_loc = J.read_loc("ui_effect_loc")
    bullets: dict[str, list[str]] = {}
    for r in J.read_tsv("ui_effect_jct"):
        ab, eff = r.get("ability", ""), r.get("effect", "")
        if not ab or not eff:
            continue
        txt = J.clean_text(resolve_loc(ui_loc, f"unit_abilities_additional_ui_effects_localised_text_{eff}"))
        if txt:
            bullets.setdefault(ab, []).append(txt)

    # 2) phase data, indexed by phase id
    phase_row = {r["id"]: r for r in J.read_tsv("phases") if r.get("id")}
    stats_by_phase: dict[str, list[str]] = {}
    for r in J.read_tsv("phase_stats"):
        ph = r.get("phase", "")
        if not ph:
            continue
        signed, _, _ = J.fmt_number(r.get("value", ""))
        stats_by_phase.setdefault(ph, []).append(f"{_clean_stat(r.get('stat',''))} {signed}")
    attrs_by_phase: dict[str, list[str]] = {}
    for r in J.read_tsv("phase_attrs"):
        ph = r.get("phase", "")
        if not ph:
            continue
        verb = "grants" if r.get("attribute_type", "") == "positive" else "removes"
        attrs_by_phase.setdefault(ph, []).append(f"{verb} {r.get('attribute','')}")

    # 3) ability -> phase clauses (with target scope)
    def scope(r) -> str:
        who = []
        if r.get("target_self") == "true":
            who.append("Self")
        if r.get("target_friends") == "true":
            who.append("Allies")
        if r.get("target_enemies") == "true":
            who.append("Enemies")
        return "/".join(who) or "Effect"

    phase_clauses: dict[str, list[str]] = {}
    for r in J.read_tsv("ab_phase_jct"):
        ab, ph = r.get("special_ability", ""), r.get("phase", "")
        if not ab or ph not in phase_row:
            continue
        p = phase_row[ph]
        bits: list[str] = []
        dmg = _num(p.get("damage_amount"))
        if dmg > 0:
            # damage_amount is per-tick, per-entity. Real output = that scaled by
            # the tick rate (1/hp_change_frequency) and the target cap
            # (max_damaged_entities) -> matches the in-game "Damage Per Second".
            freq = _num(p.get("hp_change_frequency"))
            ents = int(_num(p.get("max_damaged_entities"))) or 1
            if freq > 0:  # damage-over-time
                dps_each = dmg / freq
                if ents > 1:
                    bits.append(f"~{dps_each * ents:g} dmg/s (up to {ents} targets)")
                else:
                    bits.append(f"~{dps_each:g} dmg/s")
            else:  # instant hit
                bits.append(f"{dmg:g} damage" + (f" (up to {ents} targets)" if ents > 1 else ""))
        if _num(p.get("heal_amount")) > 0:
            bits.append(f"heals {int(_num(p['heal_amount']))}")
        if _num(p.get("barrier_heal_amount")) > 0:
            bits.append(f"+{int(_num(p['barrier_heal_amount']))} barrier")
        if p.get("resurrect") == "true":
            bits.append("resurrects/replenishes")
        if _num(p.get("execute_ratio")) > 0:
            bits.append(f"executes below {p['execute_ratio']}")
        if p.get("imbue_magical") == "true":
            bits.append("imbues magical attacks")
        if p.get("imbue_ignition") == "true":
            bits.append("imbues flaming attacks")
        bits += stats_by_phase.get(ph, [])
        bits += attrs_by_phase.get(ph, [])
        if not bits:
            continue
        dnum = _num(p.get("duration"))
        if dnum > 0:
            suffix = f" for {dnum:g}s"
        elif dnum < 0:
            suffix = " (sustained)"   # -1 = permanent aura / toggle
        else:
            suffix = ""
        clause = f"{scope(r)}: " + ", ".join(bits) + suffix
        phase_clauses.setdefault(ab, []).append(clause)

    # 4) merge bullets + phase clauses per ability
    out: dict[str, str] = {}
    for ab in set(bullets) | set(phase_clauses):
        out[ab] = _join(bullets.get(ab, []) + phase_clauses.get(ab, []))
    return out


def resolve_loc(loc: dict[str, str], key: str, _depth: int = 0) -> str:
    """Look up a loc value, following one/two levels of {{tr:other_key}} redirects."""
    val = loc.get(key, "")
    m = J.re.fullmatch(r"\{\{tr:([a-z0-9_]+)\}\}", (val or "").strip())
    if m and _depth < 3:
        return resolve_loc(loc, m.group(1), _depth + 1)
    return val


def load_abilities():
    """Return (effect_key -> [ability_key,...], ability_key -> {name,kind,mech,desc})."""
    ab_loc = J.read_loc("unit_abilities_loc")
    ab_info: dict[str, dict] = {}
    for r in J.read_tsv("unit_abilities"):
        key = r.get("key", "")
        if not key:
            continue
        ab_info[key] = {
            "name": J.clean_text(resolve_loc(ab_loc, f"unit_abilities_onscreen_name_{key}")) or key,
            "kind": r.get("source_type", ""),          # passive | active | bound | spell | ...
            "mech": r.get("type", "").replace("wh_type_", ""),  # augment / direct_damage / summon_unit ...
            "desc": J.clean_text(resolve_loc(ab_loc, f"unit_abilities_tooltip_text_{key}")),
        }
    eff_to_abilities: dict[str, list[str]] = {}
    for r in J.read_tsv("eff_ability_jct"):
        eff, ab = r.get("effect", ""), r.get("unit_ability", "")
        if not eff or not ab:
            continue
        lst = eff_to_abilities.setdefault(eff, [])
        if ab not in lst:
            lst.append(ab)
    return eff_to_abilities, ab_info


_DLC_TOKEN = J.re.compile(r"_(dlc\d+)")  # e.g. wh3_dlc27 -> dlc27


def classify_source(effect: str) -> tuple[str, str, str]:
    """(source, game, dlc_name) derived from the effect key prefix.

    source   : 'Main game' | 'DLC' | 'Core/engine'
    game     : 'WH1 (TWW1)' | 'WH2 (TWW2)' | 'WH3 (TWW3)' | ''
    dlc_name : marketing name (or raw dlc code) for DLC effects, else ''
    """
    game, dlc = J.game_and_dlc(effect)  # reuses the same prefix parsing as traits
    if game == "Unknown":
        # No wh/wh2/wh3 prefix -> generic engine effect (building_upkeep, food_cost).
        return "Core/engine", "", ""
    if _DLC_TOKEN.search(effect):
        return "DLC", game, dlc
    return "Main game", game, ""


def readable_template(effect: str, eff_loc: dict[str, str]) -> tuple[str, bool]:
    """Cleaned loc text with the numeric placeholder made visible. (text, has_loc)."""
    tmpl = eff_loc.get(f"effects_description_{effect}")
    if not tmpl:
        return "", False
    shown = tmpl.replace("%+n", "+N").replace("%n", "N").replace("%%", "%")
    return J.clean_text(shown), True


def _join(parts: list[str]) -> str:
    """Dedupe (order-preserving) and ' | '-join non-empty parts."""
    seen, out = set(), []
    for p in parts:
        p = (p or "").strip()
        if p and p not in seen:
            seen.add(p)
            out.append(p)
    return " | ".join(out)


def build():
    eff_loc = J.read_loc("effects_loc")
    eff_to_abilities, ab_info = load_abilities()
    ability_mech = load_ability_mechanics()
    ability_cost = load_ability_cost()
    rows: list[dict] = []
    for r in J.read_tsv("effects"):
        key = r.get("effect", "")
        if not key:
            continue
        readable, has_loc = readable_template(key, eff_loc)
        source, game, dlc_name = classify_source(key)
        # Ability/bound-spell this effect enables or modifies (if any).
        keys = eff_to_abilities.get(key, [])
        abils = [ab_info[a] for a in keys if a in ab_info]
        costs = [ability_cost[a] for a in keys if a in ability_cost]
        rows.append({
            "effect": key,
            "readable": readable,
            "ability": _join([a["name"] for a in abils]),
            "ability_kind": _join([a["kind"] for a in abils]),
            "ability_wom": _join([_fmt_wom(c) for c in costs]),
            "ability_cd": _join([_fmt_cd(c) for c in costs]),
            "ability_uses": _join([_fmt_uses(c) for c in costs]),
            "ability_mech": _join([ability_mech[a] for a in keys if ability_mech.get(a)]),
            "ability_desc": _join([a["desc"] for a in abils]),
            "source": source,
            "game": game,
            "dlc": dlc_name,
            "category": r.get("category", ""),
            "positive_is_good": r.get("is_positive_value_good", ""),
            "priority": r.get("priority", ""),
            "has_loc": "yes" if has_loc else "",
            "icon": r.get("icon", ""),
        })
    rows.sort(key=lambda x: (x["category"], x["effect"]))
    return rows


HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def main():
    rows = build()
    wb = Workbook()
    ws = wb.active
    ws.title = "effects"
    cols = [
        ("effect", "Effect key"),
        ("readable", "Readable (N = the value)"),
        ("ability", "Ability / bound spell"),
        ("ability_kind", "Ability type (passive/active/bound)"),
        ("ability_wom", "WoM cost (free = innate)"),
        ("ability_cd", "Cooldown"),
        ("ability_uses", "Uses per battle"),
        ("ability_mech", "What it does (mechanics: damage/heal/buffs/summons)"),
        ("ability_desc", "Ability flavor text"),
        ("source", "Source (Main game / DLC)"),
        ("game", "Game"),
        ("dlc", "DLC name"),
        ("category", "Category"),
        ("positive_is_good", "Positive is good"),
        ("priority", "Priority"),
        ("has_loc", "Has description"),
        ("icon", "Icon"),
    ]
    widths = {
        "effect": 56, "readable": 70, "ability": 30, "ability_kind": 16,
        "ability_wom": 16, "ability_cd": 10, "ability_uses": 12,
        "ability_mech": 70, "ability_desc": 60, "source": 16, "game": 12, "dlc": 26,
        "category": 14, "positive_is_good": 15, "priority": 9,
        "has_loc": 15, "icon": 30,
    }
    ws.append([c[1] for c in cols])
    for i, _ in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(c[0], "") for c in cols])
    for i, (k, _) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(k, 18)
    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_idx = [i for i, (k, _) in enumerate(cols, start=1) if k in ("readable", "ability_mech", "ability_desc")]
    for rr in range(2, ws.max_row + 1):
        for i in wrap_idx:
            ws.cell(row=rr, column=i).alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    wb.save(OUT)
    n_loc = sum(1 for r in rows if r["has_loc"])
    cats, srcs = {}, {}
    for r in rows:
        cats[r["category"]] = cats.get(r["category"], 0) + 1
        srcs[r["source"]] = srcs.get(r["source"], 0) + 1
    print(f"Wrote {OUT}")
    print(f"  effects:          {len(rows):>6}")
    print(f"  with description: {n_loc:>6}")
    n_abil = sum(1 for r in rows if r["ability"])
    print(f"  by source:        " + ", ".join(f"{s}={n}" for s, n in sorted(srcs.items())))
    print(f"  by category:      " + ", ".join(f"{c or '(blank)'}={n}" for c, n in sorted(cats.items())))
    print(f"  with ability info:{n_abil:>6}  (effects that enable/modify a bound spell or ability)")


if __name__ == "__main__":
    main()
