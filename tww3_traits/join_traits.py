#!/usr/bin/env python3
"""
join_traits.py  -  Join the raw trait tables into a filterable traits.xlsx.

Reads the .tsv files straight out of a local clone of the Shazbot/WH3-Dump repo
(https://github.com/Shazbot/WH3-Dump) and produces two sheets:

  "traits"        - one row per BASE trait id (745). This is the id you pass to
                    cm:force_add_trait(...). Columns: id, name, game, dlc,
                    number of levels, description, and rendered effects (at the
                    trait's highest level).

  "trait_levels"  - one row per trait LEVEL (979). Columns include the level
                    number and "threshold_points" - which is exactly the number
                    you pass to the console "at <trait> <points>" command / the
                    4th arg of force_add_trait.

Data sources (paths inside the cloned repo):
  db/character_traits_tables/data__.tsv           base trait ids + flags (hidden, icon)
  db/character_trait_levels_tables/data__.tsv     key(level id), level, trait(base id), threshold_points
  db/trait_level_effects_tables/data__.tsv        trait_level, effect, effect_scope, value
  text/db/character_trait_levels__.loc.tsv        onscreen_name / explanation_text / colour_text
  text/db/effects__.loc.tsv                       effects_description_<effect>  (with %+n placeholder)

Setup (once):
  git clone --depth 1 https://github.com/Shazbot/WH3-Dump.git

Run:  python3 join_traits.py                       -> writes ./traits.xlsx
      python3 join_traits.py /path/to/WH3-Dump     -> use a repo clone elsewhere
"""
from __future__ import annotations
import csv
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
# Location of the cloned repo: CLI arg > sibling ./WH3-Dump.
REPO = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "WH3-Dump")
OUT = os.path.join(HERE, "traits.xlsx")

# Logical name -> path within the repo.
FILES = {
    "character_traits":       "db/character_traits_tables/data__.tsv",
    "character_trait_levels": "db/character_trait_levels_tables/data__.tsv",
    "trait_level_effects":    "db/trait_level_effects_tables/data__.tsv",
    "trait_levels_loc":       "text/db/character_trait_levels__.loc.tsv",
    "effects_loc":            "text/db/effects__.loc.tsv",
}

# --------------------------------------------------------------------------
# Best-effort DLC display names. The game (1/2/3) and the raw dlc code are
# always derived directly from the id and are authoritative; these marketing
# names are a convenience only. Edit / extend freely - unmapped codes fall
# back to the raw code in the "dlc" column.
# --------------------------------------------------------------------------
DLC_NAMES = {
    "main": "Base game",
    # --- WARHAMMER 1 (wh_) ---
    "wh_dlc03": "The Grim & the Grave",
    "wh_dlc04": "The King & the Warlord",
    "wh_dlc05": "Realm of the Wood Elves",
    "wh_dlc07": "Bretonnia (free LL)",
    "wh_dlc08": "Norsca",
    # --- WARHAMMER 2 (wh2_) ---
    "wh2_dlc09": "Rise of the Tomb Kings",
    "wh2_dlc10": "Curse of the Vampire Coast",
    "wh2_dlc11": "The Queen & the Crone",
    "wh2_dlc12": "The Prophet & the Warlock",
    "wh2_dlc13": "The Hunter & the Beast",
    "wh2_dlc14": "The Warden & the Paunch",
    "wh2_dlc15": "The Shadow & the Blade",
    "wh2_dlc16": "The Twisted & the Twilight",
    "wh2_dlc17": "The Silence & the Fury",
    # --- WARHAMMER 3 (wh3_) ---
    "wh3_dlc20": "Champions of Chaos",
    "wh3_dlc23": "Forge of the Chaos Dwarfs",
    "wh3_dlc24": "Shadows of Change",
    "wh3_dlc25": "Thrones of Decay",
    "wh3_dlc26": "Omens of Destruction",
    # wh3_dlc27 and a few others intentionally left unmapped -> raw code shows.
}


# --------------------------------------------------------------------------
# TSV helpers (CA dumps have a "#comment" line right after the header)
# --------------------------------------------------------------------------
def _resolve(name: str) -> str:
    path = os.path.join(REPO, FILES[name])
    if not os.path.exists(path):
        raise SystemExit(
            f"Missing: {path}\n"
            f"Clone the dump first:\n"
            f"  git clone --depth 1 https://github.com/Shazbot/WH3-Dump.git\n"
            f"or pass its path:  python3 join_traits.py /path/to/WH3-Dump"
        )
    return path


def read_tsv(name: str):
    """Yield dict rows for a data__.tsv, skipping the CA #comment line."""
    with open(_resolve(name), encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        header = next(reader)
        for row in reader:
            if not row or row[0].startswith("#"):
                continue
            # pad short rows
            row += [""] * (len(header) - len(row))
            yield dict(zip(header, row))


def read_loc(name: str) -> dict[str, str]:
    """Return {loc_key: text} for a .loc.tsv (key, text, tooltip)."""
    out: dict[str, str] = {}
    with open(_resolve(name), encoding="utf-8", newline="") as f:
        for row in csv.reader(f, delimiter="\t"):
            if len(row) < 2 or row[0].startswith("#") or row[0] == "key":
                continue
            out[row[0]] = row[1]
    return out


# --------------------------------------------------------------------------
# Derivations
# --------------------------------------------------------------------------
def game_and_dlc(trait_id: str) -> tuple[str, str]:
    """('WH1 (TWW1)'|'WH2 (TWW2)'|'WH3 (TWW3)', dlc_name_or_code) from the id."""
    parts = trait_id.split("_")
    prefix = parts[0]
    game = {"wh": "WH1 (TWW1)", "wh2": "WH2 (TWW2)", "wh3": "WH3 (TWW3)"}.get(
        prefix, "Unknown"
    )
    dlc_token = parts[1] if len(parts) > 1 else ""
    # DLC key is "<prefix>_<token>" (e.g. wh2_dlc09), except "main"/"trait"/"cp1"
    full = f"{prefix}_{dlc_token}"
    dlc = DLC_NAMES.get(dlc_token) or DLC_NAMES.get(full) or dlc_token or "?"
    return game, dlc


_CA_TAG = re.compile(r"\[\[.*?\]\]")  # CA markup, e.g. [[img:icon_general]] [[/img]]


def clean_text(s: str) -> str:
    """Strip CA formatting tags and normalise whitespace/escaped newlines."""
    if not s:
        return ""
    s = _CA_TAG.sub("", s)
    s = s.replace("\\n", " ").replace("\\t", " ").replace("\n", " ").replace("\t", " ")
    s = s.replace("\\", " ")  # any stray escape backslashes left in the loc text
    return re.sub(r"\s+", " ", s).strip()


def fmt_number(value: str) -> tuple[str, str, bool]:
    """Return (signed, unsigned, is_negative) formatted from a raw '8.0000'."""
    try:
        f = float(value)
    except ValueError:
        return value, value, False
    neg = f < 0
    a = abs(f)
    a_str = str(int(a)) if a == int(a) else f"{a:g}"
    return ("-" if neg else "+") + a_str, ("-" if neg else "") + a_str, neg


def render_effect(effect: str, value: str, eff_loc: dict[str, str]) -> str:
    """Turn an (effect key, value) into readable text via the effects loc."""
    tmpl = eff_loc.get(f"effects_description_{effect}")
    signed, unsigned, _ = fmt_number(value)
    if tmpl:
        return clean_text(
            tmpl.replace("%+n", signed).replace("%n", unsigned).replace("%%", "%")
        )
    # No loc entry: fall back to the raw key so nothing is silently dropped.
    return f"[{effect}] {signed}"


# --------------------------------------------------------------------------
# Build
# --------------------------------------------------------------------------
def build():
    # loc dictionaries
    tl_loc = read_loc("trait_levels_loc")
    eff_loc = read_loc("effects_loc")

    def name_of(level_key: str) -> str:
        return clean_text(tl_loc.get(f"character_trait_levels_onscreen_name_{level_key}", ""))

    def desc_of(level_key: str) -> str:
        return clean_text(tl_loc.get(f"character_trait_levels_explanation_text_{level_key}", ""))

    def flavour_of(level_key: str) -> str:
        return clean_text(tl_loc.get(f"character_trait_levels_colour_text_{level_key}", ""))

    # effects grouped by trait level key
    effects_by_level: dict[str, list[str]] = {}
    for r in read_tsv("trait_level_effects"):
        lk = r.get("trait_level", "")
        if not lk:
            continue
        effects_by_level.setdefault(lk, []).append(
            render_effect(r.get("effect", ""), r.get("value", ""), eff_loc)
        )

    # levels grouped by base trait
    levels_by_trait: dict[str, list[dict]] = {}
    all_levels: list[dict] = []
    for r in read_tsv("character_trait_levels"):
        base = r.get("trait", "")
        level_key = r.get("key", "")
        try:
            lvl = int(r.get("level", "0") or 0)
        except ValueError:
            lvl = 0
        rec = {
            "level_key": level_key,
            "base": base,
            "level": lvl,
            "threshold_points": r.get("threshold_points", ""),
            "name": name_of(level_key),
            "description": desc_of(level_key),
            "flavour": flavour_of(level_key),
            "effects": " | ".join(effects_by_level.get(level_key, [])),
        }
        levels_by_trait.setdefault(base, []).append(rec)
        all_levels.append(rec)

    # base traits (the force_add_trait ids)
    traits: list[dict] = []
    for r in read_tsv("character_traits"):
        key = r.get("key", "")
        lvls = sorted(levels_by_trait.get(key, []), key=lambda x: x["level"])
        game, dlc = game_and_dlc(key)
        # Pick a representative level for name/effects. The dump sometimes groups
        # several level keys - all level 1 - under one base trait, only one of
        # which carries the real effects (usually the "self" key == base id).
        # So prefer: highest level number, then a level that actually has
        # effects, then the self level.
        rep = max(
            lvls,
            key=lambda l: (l["level"], bool(l["effects"]), l["level_key"] == key),
            default=None,
        )
        # description: unique explanation texts across all levels, in level order
        seen, descs = set(), []
        for l in lvls:
            d = (l["description"] or "").strip()
            if d and d not in seen:
                seen.add(d)
                descs.append(d)
        # effects: representative level, else the highest level that has any
        effects = (rep["effects"] if rep and rep["effects"]
                   else next((l["effects"] for l in reversed(lvls) if l["effects"]), ""))
        name = (rep["name"] if rep and rep["name"]
                else next((l["name"] for l in reversed(lvls) if l["name"]), ""))
        traits.append({
            "id": key,
            "name": name or "",
            "game": game,
            "dlc": dlc,
            "max_level": max((l["level"] for l in lvls), default=0),
            "description": " / ".join(descs),
            "effects_top_level": effects,
            "flavour": next((l["flavour"] for l in reversed(lvls) if l["flavour"]), "") or "",
            "hidden": r.get("hidden", ""),
            "icon": r.get("icon", ""),
        })

    traits.sort(key=lambda t: t["id"])
    all_levels.sort(key=lambda l: (l["base"], l["level"]))
    return traits, all_levels


# --------------------------------------------------------------------------
# Excel output
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_sheet(ws, columns, rows, widths, wrap_cols):
    ws.append([c[1] for c in columns])
    for i, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(c[0], "") for c in columns])
    # widths + wrapping
    for i, (keyname, _) in enumerate(columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(keyname, 18)
    if wrap_cols:
        wrap = Alignment(wrap_text=True, vertical="top")
        wrap_idx = [i for i, (k, _) in enumerate(columns, start=1) if k in wrap_cols]
        for r in range(2, ws.max_row + 1):
            for i in wrap_idx:
                ws.cell(row=r, column=i).alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"


def main():
    traits, levels = build()

    wb = Workbook()

    # Sheet 1: base traits
    ws1 = wb.active
    ws1.title = "traits"
    cols1 = [
        ("id", "Trait ID (force_add_trait)"),
        ("name", "Name"),
        ("game", "Game"),
        ("dlc", "DLC / Source"),
        ("max_level", "Max level"),
        ("description", "Description (what it does)"),
        ("effects_top_level", "Effects (highest level)"),
        ("flavour", "Flavour text"),
        ("hidden", "Hidden"),
        ("icon", "Icon"),
    ]
    widths1 = {
        "id": 46, "name": 26, "game": 12, "dlc": 26, "max_level": 9,
        "description": 48, "effects_top_level": 60, "flavour": 40,
        "hidden": 8, "icon": 16,
    }
    write_sheet(ws1, cols1, traits, widths1,
                wrap_cols={"description", "effects_top_level", "flavour"})

    # Sheet 2: per-level detail
    ws2 = wb.create_sheet("trait_levels")
    cols2 = [
        ("base", "Trait ID (force_add_trait)"),
        ("level", "Level"),
        ("threshold_points", "Threshold points (the 'at' points)"),
        ("name", "Name"),
        ("level_key", "Level id"),
        ("description", "Description"),
        ("effects", "Effects (this level)"),
        ("flavour", "Flavour text"),
    ]
    widths2 = {
        "base": 46, "level": 7, "threshold_points": 16, "name": 26,
        "level_key": 48, "description": 40, "effects": 60, "flavour": 40,
    }
    write_sheet(ws2, cols2, levels, widths2,
                wrap_cols={"description", "effects", "flavour"})

    # Sheet 3: read me
    ws3 = wb.create_sheet("README")
    notes = [
        ["TWW trait database", ""],
        ["", ""],
        ["Sheet 'traits'", "One row per BASE trait id (this is what you pass to cm:force_add_trait)."],
        ["Sheet 'trait_levels'", "One row per trait level. 'Threshold points' is the number you pass to the console 'at <trait> <points>' command (= 4th arg of force_add_trait)."],
        ["", ""],
        ["Game column", "Derived from the id prefix: wh_ = WH1, wh2_ = WH2, wh3_ = WH3. Authoritative."],
        ["DLC column", "Derived from the id (e.g. wh2_dlc09). Marketing names are best-effort; unmapped codes show the raw code. Edit DLC_NAMES in join_traits.py to extend."],
        ["Effects", "Real mechanical effects from db/trait_level_effects_tables, rendered via text/db/effects__.loc (%+n replaced by the value)."],
        ["Description", "Human text from text/db/character_trait_levels__.loc (explanation_text). Some traits have none - the Effects column then tells you what it does."],
        ["Source", "Local clone of https://github.com/Shazbot/WH3-Dump  (git pull to refresh)."],
    ]
    for row in notes:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 110
    ws3["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws3.max_row + 1):
        ws3.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  traits sheet:       {len(traits):>4} base trait ids")
    print(f"  trait_levels sheet: {len(levels):>4} trait levels")
    n_desc = sum(1 for t in traits if t["description"])
    n_eff = sum(1 for t in traits if t["effects_top_level"])
    print(f"  with description:   {n_desc:>4}")
    print(f"  with effects:       {n_eff:>4}")


if __name__ == "__main__":
    main()
