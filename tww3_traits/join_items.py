#!/usr/bin/env python3
"""
join_items.py  -  Join the raw item (ancillary) tables into a filterable items.xlsx.

Reads the .tsv files straight out of a local clone of the Shazbot/WH3-Dump repo
(https://github.com/Shazbot/WH3-Dump) and produces one big "items" sheet, one
row per item id, mirroring the twdb.io/warhammer3 Items browser filters:

  category      arcane_item / armour / enchanted_item / talisman / weapon /
                mount / general / form  (the TWDB "Category" facet)
  subcategory   no subcategory / armour_rune / banner / banner_rune /
                character_rune / engineering_rune / follower / gift / mark /
                poison / rune / spell_fragment  (the TWDB "Subcategory" facet)
  rarity        common / uncommon / rare / crafted / unique
                (from uniqueness_score via ancillary_uniqueness_groupings)
  faction       "" if usable by all, else the faction/culture it is locked to
  character     "" if any character, else the specific character(s) it is for

Data sources (paths inside the cloned repo):
  db/ancillaries_tables/data__.tsv                    item ids + flags (category, subcategory,
                                                      uniqueness_score, faction_set, legendary_item,
                                                      randomly_dropped, can_be_stolen/destroyed, ...)
  db/ancillary_to_effects_tables/data__.tsv           ancillary, effect, effect_scope, value
  db/ancillary_uniqueness_groupings_tables/data__.tsv uniqueness score range -> rarity (ui_state)
  db/ancillaries_included_agent_subtypes_tables/...   agent_subtype -> ancillary (character lock)
  db/faction_set_items_tables/data__.tsv              faction_set -> factions
  text/db/ancillaries__.loc.tsv                       onscreen_name / explanation_text / colour_text
  text/db/effects__.loc.tsv                           effects_description_<effect>  (with %+n placeholder)
  text/db/agent_subtypes__.loc.tsv                    nice character names for the "character" column

Setup (once):
  git clone --depth 1 https://github.com/Shazbot/WH3-Dump.git

Run:  python3 join_items.py                       -> writes ./items.xlsx
      python3 join_items.py /path/to/WH3-Dump     -> use a repo clone elsewhere
"""
from __future__ import annotations
import csv
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
# Location of the cloned repo: CLI arg > sibling ./WH3-Dump.
REPO = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "WH3-Dump")
OUT = os.path.join(HERE, "items.xlsx")

# Logical name -> path within the repo.
FILES = {
    "ancillaries":        "db/ancillaries_tables/data__.tsv",
    "anc_effects":        "db/ancillary_to_effects_tables/data__.tsv",
    "uniqueness":         "db/ancillary_uniqueness_groupings_tables/data__.tsv",
    "included_agents":    "db/ancillaries_included_agent_subtypes_tables/data__.tsv",
    "faction_set_items":  "db/faction_set_items_tables/data__.tsv",
    "ancillaries_loc":    "text/db/ancillaries__.loc.tsv",
    "effects_loc":        "text/db/effects__.loc.tsv",
    "agent_subtypes_loc": "text/db/agent_subtypes__.loc.tsv",
}

# Best-effort DLC display names (see join_traits.py). Authoritative info is the
# game (1/2/3) + raw dlc code from the id; these are a convenience only.
DLC_NAMES = {
    "main": "Base game",
    # --- WARHAMMER 1 (wh_) ---
    "wh_dlc03": "The Grim & the Grave",
    "wh_dlc04": "The King & the Warlord",
    "wh_dlc05": "Realm of the Wood Elves",
    "wh_dlc07": "Bretonnia (free LL)",
    "wh_dlc08": "Norsca",
    # --- WARHAMMER 2 (wh2_) ---
    "wh2_dlc09": "Rise of the Tomb Kings",
    "wh2_dlc10": "Curse of the Vampire Coast",
    "wh2_dlc11": "The Queen & the Crone",
    "wh2_dlc12": "The Prophet & the Warlock",
    "wh2_dlc13": "The Hunter & the Beast",
    "wh2_dlc14": "The Warden & the Paunch",
    "wh2_dlc15": "The Shadow & the Blade",
    "wh2_dlc16": "The Twisted & the Twilight",
    "wh2_dlc17": "The Silence & the Fury",
    # --- WARHAMMER 3 (wh3_) ---
    "wh3_dlc20": "Champions of Chaos",
    "wh3_dlc23": "Forge of the Chaos Dwarfs",
    "wh3_dlc24": "Shadows of Change",
    "wh3_dlc25": "Thrones of Decay",
    "wh3_dlc26": "Omens of Destruction",
    "wh3_dlc27": "Voyages of the Damned / Masque of Slaanesh",
}

# Human labels for the TWDB category / subcategory codes.
CATEGORY_NAMES = {
    "arcane_item": "Arcane Item",
    "armour": "Armour",
    "enchanted_item": "Enchanted Item",
    "talisman": "Talisman",
    "weapon": "Weapon",
    "mount": "Mount",
    "general": "General",
    "form": "Form",
}
SUBCATEGORY_NAMES = {
    "": "No subcategory",
    "armour_rune": "Armour Rune",
    "banner": "Banner",
    "banner_rune": "Banner Rune",
    "character_rune": "Character Rune",
    "engineering_rune": "Engineering Rune",
    "follower": "Follower",
    "gift": "Gift of Chaos",
    "mark": "Mark of Chaos",
    "poison": "Poison",
    "rune": "Rune",
    "spell_fragment": "Spell Fragment",
}


# --------------------------------------------------------------------------
# TSV helpers (CA dumps have a "#comment" line right after the header)
# --------------------------------------------------------------------------
def _resolve(name: str) -> str:
    path = os.path.join(REPO, FILES[name])
    if not os.path.exists(path):
        raise SystemExit(
            f"Missing: {path}\n"
            f"Clone the dump first:\n"
            f"  git clone --depth 1 https://github.com/Shazbot/WH3-Dump.git\n"
            f"or pass its path:  python3 join_items.py /path/to/WH3-Dump"
        )
    return path


def read_tsv(name: str):
    """Yield dict rows for a data__.tsv, skipping the CA #comment line."""
    with open(_resolve(name), encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        header = next(reader)
        for row in reader:
            if not row or row[0].startswith("#"):
                continue
            row += [""] * (len(header) - len(row))
            yield dict(zip(header, row))


def read_loc(name: str) -> dict[str, str]:
    """Return {loc_key: text} for a .loc.tsv (key, text, tooltip)."""
    out: dict[str, str] = {}
    with open(_resolve(name), encoding="utf-8", newline="") as f:
        for row in csv.reader(f, delimiter="\t"):
            if len(row) < 2 or row[0].startswith("#") or row[0] == "key":
                continue
            out[row[0]] = row[1]
    return out


# --------------------------------------------------------------------------
# Derivations
# --------------------------------------------------------------------------
def game_and_dlc(item_id: str) -> tuple[str, str]:
    """('WH1 (TWW1)'|'WH2 (TWW2)'|'WH3 (TWW3)', dlc_name_or_code) from the id."""
    parts = item_id.split("_")
    prefix = parts[0]
    game = {"wh": "WH1 (TWW1)", "wh2": "WH2 (TWW2)", "wh3": "WH3 (TWW3)"}.get(
        prefix, "Unknown"
    )
    dlc_token = parts[1] if len(parts) > 1 else ""
    full = f"{prefix}_{dlc_token}"
    dlc = DLC_NAMES.get(dlc_token) or DLC_NAMES.get(full) or dlc_token or "?"
    return game, dlc


_CA_TAG = re.compile(r"\[\[.*?\]\]")  # CA markup, e.g. [[img:icon_general]] [[/img]]


def clean_text(s: str) -> str:
    """Strip CA formatting tags and normalise whitespace/escaped newlines."""
    if not s:
        return ""
    s = _CA_TAG.sub("", s)
    s = s.replace("\\n", " ").replace("\\t", " ").replace("\n", " ").replace("\t", " ")
    s = s.replace("\\", " ")
    return re.sub(r"\s+", " ", s).strip()


def fmt_number(value: str) -> tuple[str, str, bool]:
    """Return (signed, unsigned, is_negative) formatted from a raw '8.0000'."""
    try:
        f = float(value)
    except ValueError:
        return value, value, False
    neg = f < 0
    a = abs(f)
    a_str = str(int(a)) if a == int(a) else f"{a:g}"
    return ("-" if neg else "+") + a_str, ("-" if neg else "") + a_str, neg


def render_effect(effect: str, value: str, eff_loc: dict[str, str]) -> str:
    """Turn an (effect key, value) into readable text via the effects loc."""
    tmpl = eff_loc.get(f"effects_description_{effect}")
    signed, unsigned, _ = fmt_number(value)
    if tmpl:
        return clean_text(
            tmpl.replace("%+n", signed).replace("%n", unsigned).replace("%%", "%")
        )
    return f"[{effect}] {signed}"


# Ranges from db/ancillary_uniqueness_groupings_tables (min..max -> ui_state).
# "legendary" is what the game calls score 200; twdb.io labels it "Unique".
def build_rarity(rows) -> "callable":
    ranges = []  # (lo, hi, state)
    for r in rows:
        try:
            lo = int(float(r.get("uniqueness_min", "0")))
            hi = int(float(r.get("uniqueness_max", "0")))
        except ValueError:
            continue
        state = (r.get("ui_state", "") or "").strip()
        if state == "legendary":
            state = "unique"
        ranges.append((min(lo, hi), max(lo, hi), state))

    def rarity_of(score_raw: str) -> str:
        try:
            s = int(float(score_raw))
        except (ValueError, TypeError):
            return ""
        for lo, hi, state in ranges:
            if lo <= s <= hi:
                return state
        # Outside every defined band: >=200 is a unique/legendary item, and the
        # low scores (1, <35) are the plain common pool.
        if s >= 200:
            return "unique"
        return "common"

    return rarity_of


def pretty_faction_set(fs: str) -> str:
    """Turn a faction_set code into a readable lock label ('' = usable by all)."""
    fs = (fs or "").strip()
    if fs == "" or fs == "all":
        return ""
    return (fs.replace("anc_set_exclusive_", "")
              .replace("anc_set_multi_", "")
              .replace("_", " ").strip())


# --------------------------------------------------------------------------
# Build
# --------------------------------------------------------------------------
def build():
    anc_loc = read_loc("ancillaries_loc")
    eff_loc = read_loc("effects_loc")
    agent_loc = read_loc("agent_subtypes_loc")
    rarity_of = build_rarity(read_tsv("uniqueness"))

    def name_of(key: str) -> str:
        return clean_text(anc_loc.get(f"ancillaries_onscreen_name_{key}", ""))

    def desc_of(key: str) -> str:
        return clean_text(anc_loc.get(f"ancillaries_explanation_text_{key}", ""))

    def flavour_of(key: str) -> str:
        return clean_text(anc_loc.get(f"ancillaries_colour_text_{key}", ""))

    # Generic loc names ("Legendary Lord", "Lord", "Hero") don't say WHICH
    # character, so for those we fall back to the raw subtype key (which does,
    # e.g. wh2_dlc09_tmb_arkhan -> "tmb arkhan").
    _GENERIC = {"", "legendary lord", "lord", "hero", "generic"}

    def agent_name(subtype: str) -> str:
        for pat in (f"agent_subtypes_onscreen_name_override_{subtype}",
                    f"agent_subtypes_onscreen_name_{subtype}"):
            t = clean_text(agent_loc.get(pat, ""))
            if t and t.strip().lower() not in _GENERIC:
                return t
        # Fall back to the raw key with the game/dlc prefix trimmed off.
        return re.sub(r"^wh\d?_(?:dlc\d+|main|pro\d+|cp\d+|twa\d+)_", "", subtype) or subtype

    # effects grouped by ancillary key
    effects_by_anc: dict[str, list[str]] = {}
    for r in read_tsv("anc_effects"):
        k = r.get("ancillary", "")
        if not k:
            continue
        effects_by_anc.setdefault(k, []).append(
            render_effect(r.get("effect", ""), r.get("value", ""), eff_loc)
        )

    # character lock: agent_subtype(s) that this item is included for
    chars_by_anc: dict[str, list[str]] = {}
    for r in read_tsv("included_agents"):
        anc = r.get("ancillary", "")
        sub = r.get("agent_subtype", "")
        if anc and sub:
            chars_by_anc.setdefault(anc, []).append(sub)

    items: list[dict] = []
    for r in read_tsv("ancillaries"):
        key = r.get("key", "")
        game, dlc = game_and_dlc(key)
        cat = r.get("category", "")
        sub = r.get("subcategory", "")
        subs = sorted(set(chars_by_anc.get(key, [])))
        char_names = ", ".join(agent_name(s) for s in subs)
        items.append({
            "id": key,
            "name": name_of(key) or "",
            "category": CATEGORY_NAMES.get(cat, cat or ""),
            "subcategory": SUBCATEGORY_NAMES.get(sub, sub or "No subcategory"),
            "rarity": rarity_of(r.get("uniqueness_score", "")).capitalize(),
            "game": game,
            "dlc": dlc,
            "faction_lock": pretty_faction_set(r.get("faction_set", "")),
            "character_lock": char_names,
            "effects": " | ".join(effects_by_anc.get(key, [])),
            "description": desc_of(key),
            "flavour": flavour_of(key),
            "legendary": "yes" if r.get("legendary_item", "") == "true" else "",
            "randomly_dropped": "yes" if r.get("randomly_dropped", "") == "true" else "",
            "can_be_stolen": "yes" if r.get("can_be_stolen", "") == "true" else "",
            "can_be_destroyed": "yes" if r.get("can_be_destroyed", "") == "true" else "",
        })

    items.sort(key=lambda t: (t["category"], t["id"]))
    return items


# --------------------------------------------------------------------------
# Excel output
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_sheet(ws, columns, rows, widths, wrap_cols):
    ws.append([c[1] for c in columns])
    for i, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(c[0], "") for c in columns])
    for i, (keyname, _) in enumerate(columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = widths.get(keyname, 18)
    if wrap_cols:
        wrap = Alignment(wrap_text=True, vertical="top")
        wrap_idx = [i for i, (k, _) in enumerate(columns, start=1) if k in wrap_cols]
        for r in range(2, ws.max_row + 1):
            for i in wrap_idx:
                ws.cell(row=r, column=i).alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"


def main():
    items = build()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "items"
    cols1 = [
        ("id", "Item ID"),
        ("name", "Name"),
        ("category", "Category"),
        ("subcategory", "Subcategory"),
        ("rarity", "Rarity"),
        ("game", "Game"),
        ("dlc", "DLC / Source"),
        ("faction_lock", "Faction lock (blank = any)"),
        ("character_lock", "Character lock (blank = any)"),
        ("effects", "Effects"),
        ("description", "Description (what it does)"),
        ("flavour", "Flavour text"),
        ("legendary", "Legendary"),
        ("randomly_dropped", "Random drop"),
        ("can_be_stolen", "Stealable"),
        ("can_be_destroyed", "Destroyable"),
    ]
    widths1 = {
        "id": 48, "name": 30, "category": 14, "subcategory": 16, "rarity": 11,
        "game": 12, "dlc": 26, "faction_lock": 22, "character_lock": 30,
        "effects": 66, "description": 48, "flavour": 40,
        "legendary": 10, "randomly_dropped": 11, "can_be_stolen": 10,
        "can_be_destroyed": 11,
    }
    write_sheet(ws1, cols1, items, widths1,
                wrap_cols={"effects", "description", "flavour", "character_lock"})

    # README
    ws3 = wb.create_sheet("README")
    notes = [
        ["TWW item (ancillary) database", ""],
        ["", ""],
        ["Sheet 'items'", "One row per item id (ancillary). Turn on the AutoFilter dropdowns in row 1 to filter."],
        ["", ""],
        ["Category", "arcane_item / armour / enchanted_item / talisman / weapon / mount / general / form. Same as the twdb.io Category facet."],
        ["Subcategory", "No subcategory / Armour Rune / Banner / Banner Rune / Character Rune / Engineering Rune / Follower / Gift of Chaos / Mark of Chaos / Poison / Rune / Spell Fragment."],
        ["Rarity", "common / uncommon / rare / crafted / unique. Derived from uniqueness_score via db/ancillary_uniqueness_groupings_tables (game's 'legendary' band shown as Unique, matching twdb)."],
        ["Game / DLC", "Derived from the id prefix (wh_ = WH1, wh2_ = WH2, wh3_ = WH3) + dlc code. Marketing DLC names are best-effort; unmapped codes show the raw code."],
        ["Faction lock", "Blank = usable by any faction. Otherwise the faction/culture the item is locked to (from ancillaries.faction_set; 'all' shown as blank)."],
        ["Character lock", "Blank = any character. Otherwise the specific character(s) the item is included for (from ancillaries_included_agent_subtypes)."],
        ["Effects", "Mechanical effects from db/ancillary_to_effects_tables, rendered via text/db/effects__.loc (%+n replaced by the value)."],
        ["Description / Flavour", "explanation_text / colour_text from text/db/ancillaries__.loc."],
        ["Source", "Local clone of https://github.com/Shazbot/WH3-Dump  (git pull to refresh)."],
    ]
    for row in notes:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 110
    ws3["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws3.max_row + 1):
        ws3.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws3.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  items sheet:      {len(items):>4} item ids")
    print(f"  with effects:     {sum(1 for t in items if t['effects']):>4}")
    print(f"  with description: {sum(1 for t in items if t['description']):>4}")
    print(f"  faction-locked:   {sum(1 for t in items if t['faction_lock']):>4}")
    print(f"  character-locked: {sum(1 for t in items if t['character_lock']):>4}")
    by_cat: dict[str, int] = {}
    for t in items:
        by_cat[t["category"]] = by_cat.get(t["category"], 0) + 1
    print("  by category: " + ", ".join(f"{k}={v}" for k, v in sorted(by_cat.items())))


if __name__ == "__main__":
    main()
