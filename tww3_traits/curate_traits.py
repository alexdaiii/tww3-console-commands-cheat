#!/usr/bin/env python3
"""
curate_traits.py - Build a lean "max power, base game only" trait set.

Filters the base trait ids to:
  * base-game only  (dlc token == 'main', i.e. wh_main / wh2_main / wh3_main)
  * not hidden
  * has at least one BENEFICIAL effect in Economy / Survivability / Army
  * has NO clearly harmful combat effect (so we don't add self-debuffs)

Buckets each effect into economy / survivability / army / other and prints the
result grouped, then writes power_traits.lua.
"""
from __future__ import annotations
import os
import join_traits as J   # reuse read_tsv/read_loc/render_effect/clean_text/game_and_dlc

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_LUA = os.path.join(HERE, "power_traits.lua")
OUT_XLSX = os.path.join(HERE, "power_traits.xlsx")

# --- effect classification by rendered text (lowercased) --------------------
SURVIVE_KW = [
    "armour", "physical resistance", "spell resistance", "magic resistance",
    "missile resistance", "ward save", "hit points", "barrier", "regeneration",
    "melee defence", "unbreakable", "immune to", "wounded instead of killed",
    "damage taken", "resistance to", "healing rate", "fatigue", "health",
    "leadership", "vigour", "morale",   # morale = staying power (anti-rout)
]
ECON_KW = [
    "income", "gold", "upkeep", "growth", "construction cost", "build cost",
    "research", "trade", "sacking", "loot", "post-battle",
    "tax", "money", "cost of", "replenish", "recruit", "control", "public order",
]
ATTACK_KW = [
    "melee attack", "weapon strength", "charge bonus", "armour-piercing",
    "armour piercing", "bonus vs", "magic attacks", "spell damage",
    "missile damage", "missile strength", "reload", "ammunition", "damage:",
    "attack interval", "explosion", "flaming attacks",
]
# Army = anything applied force-wide (by effect scope) OR clearly unit-wide text.
ARMY_SCOPE_HINT = "force"
ARMY_KW = ["for all", "all units", "army", "reinforc", "campaign movement", "ambush"]

# effects that are self-harmful in battle -> exclude the whole trait.
# (Each is the effect name followed by a negative value, so we never catch a buff.)
HARM_KW = [
    "melee attack: -", "leadership: -", "speed: -", "charge bonus: -",
    "weapon strength: -", "melee defence: -", "armour: -", "hit points: -",
    "resistance: -", "ward save: -", "missile strength: -",
    "bonus vs large: -", "bonus vs infantry: -",
]

# User rule: skip any trait that touches CORRUPTION (keep control / public order).
CORRUPTION_KW = "corruption"

_NEG = __import__("re").compile(r"-\s*\d")


def is_unwanted(rendered: str) -> bool:
    """Bundled downsides that shouldn't ride along on a power set -> drop trait."""
    low = rendered.lower()
    has_neg = bool(_NEG.search(rendered))
    if "plague" in low:                                   # +plague spread / duration
        return True
    if ("public order" in low or "public_order" in low) and has_neg:   # -public order
        return True
    if "diplomatic relations" in low and has_neg:         # -diplo (e.g. -20 with Empire)
        return True
    if ("winds of magic power reserve" in low and has_neg
            and "enemy" not in low and "capacity" not in low):         # -X per turn drain
        return True
    return False


def bucket(text: str, scope: str) -> str:
    t = text.lower()
    sc = (scope or "").lower()
    # enemy debuffs, and anything applied force-wide, are ARMY effects (they help
    # the whole stack) - decide this before the survivability keywords so an
    # "Enemy leadership: -5" or a force-wide leadership aura isn't mislabelled.
    if "enemy" in t or ARMY_SCOPE_HINT in sc or "aura" in t:
        return "army"
    if any(k in t for k in SURVIVE_KW):
        return "survive"
    if any(k in t for k in ATTACK_KW):
        return "attack"
    if ARMY_SCOPE_HINT in (scope or "").lower() or any(k in t for k in ARMY_KW):
        return "army"
    if any(k in t for k in ECON_KW):
        return "econ"
    return "other"


def main():
    tl_loc = J.read_loc("trait_levels_loc")
    eff_loc = J.read_loc("effects_loc")

    def name_of(k):
        return J.clean_text(tl_loc.get(f"character_trait_levels_onscreen_name_{k}", ""))

    # effects (rendered + scope) grouped by trait level key
    raw_by_level: dict[str, list[tuple[str, str, str]]] = {}
    for r in J.read_tsv("trait_level_effects"):
        lk = r.get("trait_level", "")
        if not lk:
            continue
        rendered = J.render_effect(r.get("effect", ""), r.get("value", ""), eff_loc)
        raw_by_level.setdefault(lk, []).append(
            (rendered, r.get("effect_scope", ""), r.get("value", ""))
        )

    # levels grouped by base trait -> pick representative (highest, has effects, self)
    levels_by_trait: dict[str, list[dict]] = {}
    for r in J.read_tsv("character_trait_levels"):
        base = r.get("trait", "")
        lk = r.get("key", "")
        try:
            lvl = int(r.get("level", "0") or 0)
        except ValueError:
            lvl = 0
        levels_by_trait.setdefault(base, []).append(
            {"lk": lk, "lvl": lvl, "eff": raw_by_level.get(lk, [])}
        )

    rows = []
    for r in J.read_tsv("character_traits"):
        key = r.get("key", "")
        parts = key.split("_")
        dlc_token = parts[1] if len(parts) > 1 else ""
        if dlc_token != "main":            # base game only
            continue
        if (r.get("hidden", "") or "").lower() == "true":
            continue
        # skip only internal placeholder traits (…_dummy_…). Prologue traits are
        # KEPT: force_add_trait applies them in the main campaign too (e.g.
        # "Saviour", the enemy-leadership auras) even though they're tutorial-tagged.
        if "dummy" in key:
            continue
        # NB: "defeated X" trophy traits are NOT blanket-excluded; the penalty
        # filters below (negative diplomacy / public order / stats, corruption,
        # plague) drop the ones that carry a downside and keep the clean ones.
        lvls = levels_by_trait.get(key, [])
        rep = max(lvls, key=lambda l: (l["lvl"], bool(l["eff"]), l["lk"] == key),
                  default=None)
        if not rep or not rep["eff"]:
            continue
        max_lvl = max((l["lvl"] for l in lvls), default=1)
        buckets = {"econ": [], "survive": [], "army": [], "attack": [], "other": []}
        harmful = False
        touches_corruption = False
        bundled_bad = False
        for rendered, scope, _val in rep["eff"]:
            low = rendered.lower()
            if CORRUPTION_KW in low:
                touches_corruption = True
            # "Enemy leadership: -X" (and other enemy debuffs) are BUFFS, not
            # self-harm - never let the negative-stat filter catch them.
            if "enemy" not in low and any(h in low for h in HARM_KW):
                harmful = True
            if is_unwanted(rendered):
                bundled_bad = True
            buckets[bucket(rendered, scope)].append(rendered)
        # drop self-debuffs, corruption traits, and traits with bundled downsides
        if harmful or touches_corruption or bundled_bad:
            continue
        # must contribute to at least one of the four target areas
        if not (buckets["econ"] or buckets["survive"] or buckets["army"] or buckets["attack"]):
            continue
        rows.append({
            "id": key,
            "name": name_of(rep["lk"]) or name_of(key),
            "points": max(max_lvl, 1),
            "econ": buckets["econ"],
            "survive": buckets["survive"],
            "army": buckets["army"],
            "attack": buckets["attack"],
            "other": buckets["other"],
        })

    rows.sort(key=lambda x: x["id"])

    CATS = ("Survivability", "Attack", "Army", "Economy")
    BKEY = {"Survivability": "survive", "Attack": "attack",
            "Army": "army", "Economy": "econ"}

    # primary category = the target bucket with the most effects
    def primary(x):
        counts = {c: len(x[BKEY[c]]) for c in CATS}
        return max(counts, key=counts.get) if any(counts.values()) else "Economy"

    grouped: dict[str, list[dict]] = {c: [] for c in CATS}
    for x in rows:
        grouped[primary(x)].append(x)

    for cat in CATS:
        print(f"\n{'='*70}\n{cat}  ({len(grouped[cat])} traits)\n{'='*70}")
        for x in grouped[cat]:
            eff = x["survive"] + x["attack"] + x["army"] + x["econ"]
            print(f"  {x['name'] or x['id']}  [{x['id']}]")
            for e in eff[:6]:
                print(f"      + {e}")

    # write the lua
    lua_entries = []
    for cat in CATS:
        lua_entries.append(f"    -- ---- {cat} ----")
        for x in grouped[cat]:
            nm = (x["name"] or x["id"]).replace("--", "-")
            lua_entries.append(f'    {{ "{x["id"]}", {x["points"]} }},  -- {nm}')
    body = "\n".join(lua_entries)

    with open(OUT_LUA, "w", encoding="utf-8") as f:
        f.write(LUA.format(count=len(rows), body=body))

    write_xlsx(rows, grouped, CATS, BKEY)

    print(f"\nWrote {OUT_LUA}")
    print(f"Wrote {OUT_XLSX}")
    print(f"  {len(rows)} traits: "
          + ", ".join(f"{len(grouped[c])} {c.lower()}" for c in CATS))


# --------------------------------------------------------------------------
# Excel output
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(bold=True, color="FFFFFF")
# soft category tints
CAT_FILL = {
    "Survivability": PatternFill("solid", fgColor="DDEBF7"),
    "Attack":        PatternFill("solid", fgColor="FCE4D6"),
    "Army":          PatternFill("solid", fgColor="E2EFDA"),
    "Economy":       PatternFill("solid", fgColor="FFF2CC"),
}


def _is_conditional(effs) -> bool:
    """True if any effect only applies in a narrow case (unit/faction/situation)."""
    for e in effs:
        low = e.lower()
        if (" for " in low and "unit" in low) or " when " in low or " against " in low:
            return True
    return False


def write_xlsx(rows, grouped, CATS, BKEY):
    wb = Workbook()
    ws = wb.active
    ws.title = "power_traits"
    columns = [
        ("category", "Category"),
        ("id", "Trait ID (force_add_trait)"),
        ("name", "Name"),
        ("points", "Points"),
        ("game", "Game"),
        ("conditional", "Conditional?"),
        ("survive", "Survivability effects"),
        ("attack", "Attack effects"),
        ("army", "Army effects"),
        ("econ", "Economy effects"),
        ("other", "Other effects"),
    ]
    ws.append([c[1] for c in columns])
    for i in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for cat in CATS:
        for x in grouped[cat]:
            all_eff = x["survive"] + x["attack"] + x["army"] + x["econ"]
            rec = {
                "category": cat,
                "id": x["id"],
                "name": x["name"] or x["id"],
                "points": x["points"],
                "game": J.game_and_dlc(x["id"])[0],
                "conditional": "yes" if _is_conditional(all_eff) else "",
                "survive": " | ".join(x["survive"]),
                "attack": " | ".join(x["attack"]),
                "army": " | ".join(x["army"]),
                "econ": " | ".join(x["econ"]),
                "other": " | ".join(x["other"]),
            }
            ws.append([rec[c[0]] for c in columns])
            ws.cell(row=ws.max_row, column=1).fill = CAT_FILL.get(cat)

    widths = {
        "category": 15, "id": 46, "name": 26, "points": 8, "game": 12,
        "conditional": 12, "survive": 42, "attack": 42, "army": 46,
        "econ": 42, "other": 34,
    }
    for i, (k, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(k, 18)
    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_cols = {"survive", "attack", "army", "econ", "other", "name"}
    wrap_idx = [i for i, (k, _) in enumerate(columns, start=1) if k in wrap_cols]
    for r in range(2, ws.max_row + 1):
        for i in wrap_idx:
            ws.cell(row=r, column=i).alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    # README sheet
    ws2 = wb.create_sheet("README")
    notes = [
        ["Curated 'power' traits", ""],
        ["", ""],
        ["What this is", "Base-game (wh_main / wh2_main / wh3_main) traits with only BENEFICIAL Economy / Survivability / Army / Attack effects. Filterable - use the row-1 dropdowns."],
        ["Excluded", "DLC-locked traits, hidden traits, personality/flavour filler, self-debuffs (-melee attack etc.), and anything that modifies CORRUPTION (good/bad depends on faction)."],
        ["Category", "The trait's dominant bucket (most effects). Filter here to see just Survivability, Attack, Army, or Economy traits."],
        ["Conditional?", "'yes' = at least one effect only applies to specific units, or 'when fighting X' / 'against X'. Blank = flat, always-on, any-faction effects."],
        ["Points", "The number passed to force_add_trait / the console 'at <trait> <points>' (the trait's max level)."],
        ["Effect columns", "Effects split by bucket, rendered from the game data. A trait can have effects in more than one column."],
        ["Lua", "power_traits.lua applies this whole set. Generated by the same script (curate_traits.py)."],
        ["Source", "Local clone of https://github.com/Shazbot/WH3-Dump."],
    ]
    for row in notes:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 110
    ws2["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws2.max_row + 1):
        ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=r, column=1).font = Font(bold=True)

    wb.save(OUT_XLSX)


LUA = '''--------------------------------------------------------------------------------
-- power_traits.lua  -  lean "base game only" power set ({count} traits)
--
-- Base-game traits (wh_main / wh2_main / wh3_main) with beneficial Economy,
-- Survivability, or Army effects. No DLC-locked traits, no personality/flavour
-- filler, no self-debuffs. Run via the console "arbitrary lua" mod.
--------------------------------------------------------------------------------

local APPLY_TO = "character"   -- "character" (selected) or "faction" (leader)
local SHOW_MESSAGE = true

local TRAITS = {{
{body}
}}

local cm = get_cm and get_cm() or cm

local function get_selected_character()
    local ui = cm:get_campaign_ui_manager()
    local cqi = ui and ui:get_char_selected_cqi()
    if cqi and cqi ~= 0 then return cm:get_character_by_cqi(cqi) end
    return nil
end

local function get_faction_leader()
    local faction = cm:get_local_faction(true) or cm:get_local_faction()
    if not faction or faction:is_null_interface() then return nil end
    local leader = faction:faction_leader()
    if leader and not leader:is_null_interface() then return leader end
    return nil
end

local function run()
    local target = (APPLY_TO == "faction") and get_faction_leader() or get_selected_character()
    if not target or target:is_null_interface() then
        out("power_traits: ERROR - no valid target (" .. APPLY_TO .. "). Select a character first.")
        return
    end
    local lookup = cm:char_lookup_str(target)
    for _, entry in ipairs(TRAITS) do
        local id, pts = entry[1], entry[2] or 0
        if id and id ~= "" then
            if pts and pts > 0 then
                cm:force_add_trait(lookup, id, SHOW_MESSAGE, pts)
            else
                cm:force_add_trait(lookup, id, SHOW_MESSAGE)
            end
        end
    end
    out("power_traits: done.")
end

run()
'''

if __name__ == "__main__":
    main()
