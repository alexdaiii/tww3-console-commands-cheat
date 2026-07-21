#!/usr/bin/env python3
"""
make_apply_lua.py  -  Turn highlighted rows in traits.xlsx into an apply-traits
                      Lua script, and print the combined (stacked) effects.

Workflow
--------
1. Open traits.xlsx, highlight the rows/cells of the traits you want (any fill
   colour - e.g. green). Highlighting ANY cell in a row tags that row. Save.

2. See which colours you used:
       python3 make_apply_lua.py
   -> prints every fill colour found, how many traits have it, and examples.

3. Generate the Lua + effect summary for one colour (by number, hex, or name):
       python3 make_apply_lua.py green
       python3 make_apply_lua.py 00B050
       python3 make_apply_lua.py 1
   -> writes apply_selected_traits.lua and prints the combined effects.

Options:
  --xlsx PATH     spreadsheet to read           (default: ./traits.xlsx)
  --sheet NAME    sheet to read                 (default: traits)
  --out PATH      Lua file to write             (default: ./apply_selected_traits.lua)
  --apply-to      character | faction           (default: character)
  --base-level    apply every trait at level 1 instead of its Max level

The effect summary sums effects that stack (e.g. two "+5% Range" traits -> +10%).
It reads the effect text straight from the "Effects (highest level)" column, so
it needs nothing but the xlsx.
"""
from __future__ import annotations
import argparse
import os
import re
import sys
from collections import OrderedDict

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))


# --------------------------------------------------------------------------
# Colour handling
# --------------------------------------------------------------------------
def cell_fill(cell) -> str | None:
    """Return a normalised colour id for a cell's fill, or None if unfilled."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    c = fill.fgColor
    if c is None:
        return None
    if c.type == "rgb" and c.rgb:
        rgb = c.rgb
        if len(rgb) == 8:          # strip AA from AARRGGBB
            rgb = rgb[2:]
        if rgb == "000000" and (fill.fgColor.rgb in (None, "00000000")):
            return None
        return rgb.upper()
    if c.type == "theme" and c.theme is not None:
        return f"theme{c.theme}:{round(float(c.tint or 0), 2)}"
    if c.type == "indexed" and c.indexed is not None:
        return f"indexed{c.indexed}"
    return None


def hue_name(color: str) -> str:
    """Friendly label (green/red/...) for an RRGGBB hex; '' for non-rgb ids."""
    if not re.fullmatch(r"[0-9A-F]{6}", color):
        return ""
    r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    mx, mn = max(r, g, b), min(r, g, b)
    if mx - mn < 25:
        if mx > 220:
            return "white"
        if mx < 60:
            return "black"
        return "grey"
    if r >= g and r >= b:
        return "yellow" if g > 150 and b < 120 else ("orange" if g > 90 else "red")
    if g >= r and g >= b:
        return "green"
    return "cyan" if g > 150 else "blue"


# --------------------------------------------------------------------------
# Read the sheet
# --------------------------------------------------------------------------
def find_col(headers: list[str], *needles: str) -> int:
    for i, h in enumerate(headers):
        low = (h or "").lower()
        if all(n.lower() in low for n in needles):
            return i
    raise SystemExit(f"Could not find a column matching {needles} in: {headers}")


def read_rows(xlsx: str, sheet: str):
    wb = load_workbook(xlsx)  # default load keeps cell styles/fills
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet}' not in {xlsx}. Sheets: {wb.sheetnames}")
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    ci_id = find_col(headers, "trait id")
    ci_name = find_col(headers, "name")
    ci_eff = find_col(headers, "effects")
    try:
        ci_lvl = find_col(headers, "max level")
    except SystemExit:
        ci_lvl = None

    rows = []
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c) for c in range(1, len(headers) + 1)]
        colours = {cf for cf in (cell_fill(c) for c in cells) if cf}
        tid = cells[ci_id].value
        if not tid:
            continue
        rows.append({
            "id": str(tid),
            "name": cells[ci_name].value or "",
            "effects": cells[ci_eff].value or "",
            "max_level": (cells[ci_lvl].value if ci_lvl is not None else 1) or 1,
            "colours": colours,
        })
    return rows


# --------------------------------------------------------------------------
# Effect stacking
# --------------------------------------------------------------------------
NUM = re.compile(r"[+-]?\d+(?:\.\d+)?")


def fmt_signed(v: float) -> str:
    a = abs(v)
    a_str = str(int(a)) if a == int(a) else f"{a:g}"
    return ("-" if v < 0 else "+") + a_str


def combine_effects(selected: list[dict]):
    """Sum stacking numeric effects; count repeated non-numeric ones."""
    numeric: "OrderedDict[str, float]" = OrderedDict()  # signature -> summed value
    flat: "OrderedDict[str, int]" = OrderedDict()        # text -> count
    for row in selected:
        for piece in (p.strip() for p in str(row["effects"]).split("|")):
            if not piece:
                continue
            m = NUM.search(piece)
            if m:
                val = float(m.group())
                sig = piece[:m.start()] + "{n}" + piece[m.end():]
                numeric[sig] = numeric.get(sig, 0.0) + val
            else:
                flat[piece] = flat.get(piece, 0) + 1
    lines = [sig.replace("{n}", fmt_signed(val)) for sig, val in numeric.items()]
    for text, n in flat.items():
        lines.append(text + (f"  (x{n})" if n > 1 else ""))
    return lines


# --------------------------------------------------------------------------
# Lua generation
# --------------------------------------------------------------------------
LUA_TEMPLATE = '''--------------------------------------------------------------------------------
-- {out_name}
-- Auto-generated by make_apply_lua.py from highlighted rows in traits.xlsx.
-- Colour: {colour_label}    Traits: {n}
--
-- Run via a TWW3 "run arbitrary lua" console command. Applies the traits below
-- to the currently SELECTED character (or the faction leader if APPLY_TO="faction").
--
-- Combined effects of this set (stacking summed):
{effects_comment}
--------------------------------------------------------------------------------

local APPLY_TO    = "{apply_to}"   -- "character" (selected) or "faction" (leader)
local SHOW_MESSAGE = true

-- {{ trait_id, points }}  (points = the console "at <trait> <points>" value)
local TRAITS = {{
{trait_lines}
}}

local cm = get_cm and get_cm() or cm

local function get_selected_character()
    local ui = cm:get_campaign_ui_manager()
    local cqi = ui and ui:get_char_selected_cqi()
    if cqi and cqi ~= 0 then return cm:get_character_by_cqi(cqi) end
    return nil
end

local function get_faction_leader()
    local faction = cm:get_local_faction(true) or cm:get_local_faction()
    if not faction or faction:is_null_interface() then return nil end
    local leader = faction:faction_leader()
    if leader and not leader:is_null_interface() then return leader end
    return nil
end

local function apply_trait(character, trait_id, points)
    local lookup = cm:char_lookup_str(character)
    if points and points > 0 then
        cm:force_add_trait(lookup, trait_id, SHOW_MESSAGE, points)
    else
        cm:force_add_trait(lookup, trait_id, SHOW_MESSAGE)
    end
    out("apply_selected_traits: added '" .. tostring(trait_id) ..
        "' (points " .. tostring(points or 0) .. ") to cqi " ..
        tostring(character:command_queue_index()))
end

local function run()
    local target = (APPLY_TO == "faction") and get_faction_leader() or get_selected_character()
    if not target or target:is_null_interface() then
        out("apply_selected_traits: ERROR - no valid target (" .. APPLY_TO .. "). Select a character first.")
        return
    end
    for _, entry in ipairs(TRAITS) do
        if entry[1] and entry[1] ~= "" then apply_trait(target, entry[1], entry[2] or 0) end
    end
    out("apply_selected_traits: done (" .. tostring(#TRAITS) .. " traits).")
end

run()
'''


def build_lua(selected, out_name, colour_label, apply_to, base_level, effect_lines):
    trait_lines = []
    for row in selected:
        pts = 0 if base_level else int(row["max_level"] or 1)
        name = str(row["name"]).replace("\n", " ") if row["name"] else ""
        comment = f"  -- {name}" if name else ""
        trait_lines.append(f'    {{ "{row["id"]}", {pts} }},{comment}')
    effects_comment = "\n".join(f"--   * {ln}" for ln in effect_lines) or "--   (none)"
    return LUA_TEMPLATE.format(
        out_name=out_name,
        colour_label=colour_label,
        n=len(selected),
        apply_to=apply_to,
        effects_comment=effects_comment,
        trait_lines="\n".join(trait_lines),
    )


# --------------------------------------------------------------------------
# Palette / selection
# --------------------------------------------------------------------------
def build_palette(rows):
    """Ordered list of (colour, [rows]) for colours that tag >=1 row."""
    palette: "OrderedDict[str, list]" = OrderedDict()
    for row in rows:
        for col in row["colours"]:
            palette.setdefault(col, []).append(row)
    # sort by count desc for stable, useful display
    return sorted(palette.items(), key=lambda kv: -len(kv[1]))


def print_palette(palette):
    if not palette:
        print("No highlighted cells found. Colour some rows in the xlsx and save, then rerun.")
        return
    print("Fill colours found (highlight ANY cell in a row to tag it):\n")
    for i, (col, rws) in enumerate(palette, 1):
        label = hue_name(col)
        disp = f"#{col}" if re.fullmatch(r"[0-9A-F]{6}", col) else col
        examples = ", ".join(r["name"] or r["id"] for r in rws[:3])
        print(f"  [{i}]  {disp:<10} {label:<7} {len(rws):>3} traits   e.g. {examples}")
    print("\nGenerate for a colour:  python3 make_apply_lua.py <number|hex|name>")


def resolve_colour(arg: str, palette):
    """Map a user arg (index / hex / hue-name) to a colour id in the palette."""
    colours = [c for c, _ in palette]
    # index
    if arg.isdigit():
        idx = int(arg)
        if 1 <= idx <= len(colours):
            return colours[idx - 1]
    a = arg.lstrip("#").upper()
    if len(a) == 8:
        a = a[2:]
    if a in colours:                                   # exact hex
        return a
    matches = [c for c in colours if hue_name(c) == arg.lower()]   # hue name
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        opts = ", ".join(f"#{m}" for m in matches)
        raise SystemExit(f"'{arg}' matches several colours: {opts}. Use the exact hex or number.")
    raise SystemExit(f"Colour '{arg}' not found. Run with no args to list the palette.")


# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Highlighted traits -> apply-traits Lua + effect summary.")
    ap.add_argument("colour", nargs="?", help="Colour to select: number, hex (00B050) or name (green).")
    ap.add_argument("--xlsx", default=os.path.join(HERE, "traits-hl.xlsx"))
    ap.add_argument("--sheet", default="traits")
    ap.add_argument("--out", default=os.path.join(HERE, "apply_selected_traits.lua"))
    ap.add_argument("--apply-to", choices=["character", "faction"], default="character")
    ap.add_argument("--base-level", action="store_true", help="Apply each trait at level 1, not its Max level.")
    args = ap.parse_args()

    rows = read_rows(args.xlsx, args.sheet)
    palette = build_palette(rows)

    if not args.colour:
        print_palette(palette)
        return

    colour = resolve_colour(args.colour, palette)
    selected = dict(palette)[colour]
    label = f"#{colour} ({hue_name(colour)})" if re.fullmatch(r"[0-9A-F]{6}", colour) else colour

    effect_lines = combine_effects(selected)
    lua = build_lua(selected, os.path.basename(args.out), label,
                    args.apply_to, args.base_level, effect_lines)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(lua)

    print(f"Selected colour {label}: {len(selected)} traits\n")
    for row in selected:
        pts = 1 if args.base_level else (row["max_level"] or 1)
        print(f"  - {row['id']}   (points {pts})   {row['name']}")
    print("\nCombined effects (stacking summed):")
    for ln in effect_lines:
        print(f"  * {ln}")
    print(f"\nWrote Lua -> {args.out}")


if __name__ == "__main__":
    main()
